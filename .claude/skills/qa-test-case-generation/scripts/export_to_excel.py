import sys
import json
import os
import subprocess

def install_and_import(package):
    try:
        __import__(package)
    except ImportError:
        print(f"Installing {package}...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Ensure openpyxl is installed
install_and_import('openpyxl')
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def export_test_cases_to_excel(json_filepath, excel_filepath):
    # Load JSON data
    try:
        with open(json_filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Error reading JSON file {json_filepath}: {e}")
        sys.exit(1)
        
    test_cases = data.get('test_cases', [])
    if not test_cases:
        print("No test cases found in JSON data.")
        sys.exit(1)

    # Create workbook and select active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Define headers
    headers = ["Test Case ID", "Type", "Title / Description", "Preconditions", "Steps to Reproduce", "Expected Result"]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data
    for row_num, tc in enumerate(test_cases, 2):
        ws.cell(row=row_num, column=1, value=tc.get('id', ''))
        ws.cell(row=row_num, column=2, value=tc.get('type', ''))
        ws.cell(row=row_num, column=3, value=tc.get('title', ''))
        ws.cell(row=row_num, column=4, value=tc.get('preconditions', ''))
        
        # Format steps and expected results to handle newlines
        steps = tc.get('steps', [])
        if isinstance(steps, list):
            steps_str = "\n".join([f"{i+1}. {step}" for i, step in enumerate(steps)])
        else:
            steps_str = steps
        ws.cell(row=row_num, column=5, value=steps_str)
        
        ws.cell(row=row_num, column=6, value=tc.get('expected_result', ''))
        
        # Apply word wrap and alignment for data rows
        for col_num in range(1, 7):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Adjust column widths
    column_widths = {
        'A': 15,  # ID
        'B': 15,  # Type
        'C': 30,  # Title
        'D': 25,  # Preconditions
        'E': 45,  # Steps
        'F': 35   # Expected Result
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save workbook
    try:
        wb.save(excel_filepath)
        print(f"Successfully exported {len(test_cases)} test cases to {excel_filepath}")
    except Exception as e:
        print(f"Error saving Excel file {excel_filepath}: {e}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python export_to_excel.py <input_json_file> <output_excel_file>")
        sys.exit(1)
        
    input_json = sys.argv[1]
    output_excel = sys.argv[2]
    export_test_cases_to_excel(input_json, output_excel)
