import json
import argparse
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    exit(1)

def generate_checklist(json_path, output_path):
    if not os.path.exists(json_path):
        print(f"Error: File not found at {json_path}")
        return

    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Error parsing JSON: {e}")
        return

    list_config = data.get("listConfig", [])
    if not list_config:
        print("No 'listConfig' found in JSON.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Ads Checklist"

    # Define headers
    headers = [
        "STT", "Config Name", "Ad Type", "Mặc định (isOn)", 
        "Kịch bản Preload", "Pass: Load", "Pass: Show", 
        "Pass: Click", "Pass: Fallback (Error)", "Ghi chú Test"
    ]

    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = thin_border

    # Fills for different types
    fill_native = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_interstitial = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_banner = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_off = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Grey for isOn=False

    # Write data
    row_idx = 2
    for idx, config in enumerate(list_config, 1):
        name = config.get("configName", "N/A")
        ad_type = config.get("type", "N/A")
        is_on = config.get("isOn", False)
        is_preload = config.get("isPreloadAfterShow", False)

        # Determine row color
        row_fill = None
        if not is_on:
            row_fill = fill_off
        elif "native" in ad_type.lower():
            row_fill = fill_native
        elif "interstitial" in ad_type.lower() or "open" in ad_type.lower():
            row_fill = fill_interstitial
        elif "banner" in ad_type.lower():
            row_fill = fill_banner

        # Preload note
        preload_note = "Có Preload cho ad tiếp theo" if is_preload else "Không"
        
        # Test Note suggestions based on type
        note = ""
        if "native" in ad_type.lower():
            note = f"Check UI đè layout? Check CTA color: {config.get('ctaGradientListColor', 'Default')}"
        elif "interstitial" in ad_type.lower():
            delay = config.get("timeDelayShowInter", 0)
            note = f"Check Delay {delay}s. Bấm (x) app không bị đơ."
        elif "open_app" in ad_type.lower():
            note = "Đưa app xuống background, mở lên lại xem có hiện không."

        row_data = [
            idx, name, ad_type, "BẬT" if is_on else "TẮT",
            preload_note, "", "", "", "", note
        ]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = align_center if col_idx not in [2, 10] else align_left
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
        
        row_idx += 1

    # Adjust column widths
    column_widths = {
        'A': 5,  'B': 25, 'C': 20, 'D': 15, 'E': 20,
        'F': 12, 'G': 12, 'H': 12, 'I': 15, 'J': 40
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    try:
        wb.save(output_path)
        print(f"Da tao file checklist thanh cong tai: {output_path}")
    except Exception as e:
        print(f"Loi khi luu file Excel: {e}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Tạo Test Checklist từ file JSON config quảng cáo.')
    parser.add_argument('--input', type=str, required=True, help='Đường dẫn tới file JSON (vd: config_show_ads.json)')
    parser.add_argument('--output', type=str, default='test_ads_checklist_auto.xlsx', help='Đường dẫn xuất file Excel')
    args = parser.parse_args()

    generate_checklist(args.input, args.output)
