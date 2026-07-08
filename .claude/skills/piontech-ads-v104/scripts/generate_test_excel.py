import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def generate_excel():
    # Khởi tạo workbook mới
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ads Checklist"

    # Định nghĩa tiêu đề cột
    headers = [
        "STT", 
        "Config Name (Vị trí)", 
        "Space Name (Ad Unit mapping)", 
        "Ad Type", 
        "Ad Unit ID", 
        "Mô tả Preload (Luồng trước đó tải ads này)",
        "Pass: Load",
        "Pass: Show",
        "Pass: Click",
        "Pass: Error Handle"
    ]
    
    # Ghi header
    ws.append(headers)
    
    # Style cho header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Tùy chỉnh độ rộng cột
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 25 if col_num in [2, 3, 5, 6] else 15

    # Dữ liệu test mẫu dựa trên file tham chiếu
    # Danh sách này là một phần nhỏ để làm mẫu, có thể thêm đầy đủ sau
    data = [
        ["1", "Splash_openad3", "Splash_openad3", "Open Ad", "ca-app-pub-...", "Không preload (App khởi tạo)", "", "", "", ""],
        ["2", "Language1.1", "Language1.1_native1", "Native", "ca-app-pub-...", "Preload từ SplashFragment", "", "", "", ""],
        ["3", "Onboard3", "onboard3_native", "Native", "ca-app-pub-...", "Preload từ LanguageFragment", "", "", "", ""],
        ["4", "home_bottom", "home-bottom_adaptive", "Banner Adapt", "ca-app-pub-...", "Preload từ Onboard3", "", "", "", ""],
        ["5", "exitapp", "exitapp_native1", "Native", "ca-app-pub-...", "Preload từ HomeFragment", "", "", "", ""],
        ["6", "phone_search-search", "-", "Interstitial", "ca-app-pub-...", "Không preload", "", "", "", ""],
        ["7", "phone_search_bottom", "phone_search_1ID_adaptive", "Banner Adapt", "ca-app-pub-...", "Preload từ lúc tap Menu Home", "", "", "", ""],
        ["8", "live_search-more", "-", "Interstitial", "ca-app-pub-...", "Không preload", "", "", "", ""],
        ["9", "appresume", "appresume_openad1", "Open Ad", "ca-app-pub-...", "Khởi tạo từ HomeFragment", "", "", "", ""],
        ["10", "360_search_bettween", "view_360_bettween_1ID_native", "Native", "ca-app-pub-...", "Load trực tiếp khi scroll list", "", "", "", ""]
    ]

    # Ghi dữ liệu vào sheet
    for row_data in data:
        ws.append(row_data)

    # Căn giữa và bọc chữ
    for row in ws.iter_rows(min_row=2, max_row=len(data) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Lưu file
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_file = os.path.join(output_dir, "test_ads_checklist.xlsx")
    wb.save(output_file)
    print(f"Successfully created test checklist at: {output_file}")

if __name__ == "__main__":
    generate_excel()
