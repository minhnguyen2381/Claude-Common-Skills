import sys
import subprocess
import os
import csv

def install_and_import(package):
    try:
        __import__(package)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package])

install_and_import('openpyxl')

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import FormulaRule

def generate_excel(csv_path, output_path):
    # Read CSV
    data_by_config = {}
    headers = []
    
    with open(csv_path, mode='r', encoding='utf-8') as f:
        reader = csv.reader(f)
        try:
            headers = next(reader)
        except StopIteration:
            print("CSV file is empty.")
            return

        for row in reader:
            if len(row) < 8:
                row.extend([""] * (8 - len(row)))
            
            # row structure: Testcase_ID, Config Name, Ad Format / Placement, Category, Test Condition, Expected Result, Status, Notes/Actual
            config_name = row[1]
            if not config_name:
                config_name = "Default"
                
            if config_name not in data_by_config:
                data_by_config[config_name] = []
            
            data_by_config[config_name].append(row)

    wb = Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Define fills for conditional formatting
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Green
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red
    untested_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Yellow
    skip_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") # Gray
    
    for sheet_name, testcases in data_by_config.items():
        # Sheet name max length is 31, remove invalid chars
        safe_sheet_name = str(sheet_name)[:31].replace(":", "").replace("/", "").replace("\\", "").replace("?", "").replace("*", "").replace("[", "").replace("]", "")
        if not safe_sheet_name:
            safe_sheet_name = "Sheet"
            
        ws = wb.create_sheet(title=safe_sheet_name)
        ws.append(headers)
        
        # Format Header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        
        # Append Data
        for row in testcases:
            ws.append(row)
            
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            if adjusted_width > 50:
                adjusted_width = 50
            ws.column_dimensions[column].width = adjusted_width

        max_row = len(testcases) + 1
        
        # Add Dropdown Data Validation to Status column (G is the 7th column)
        dv = DataValidation(type="list", formula1='"Pass,Fail,Untested,Skip (N/A)"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"G2:G{max_row}")

        # Add Conditional Formatting to the entire row based on Status column (G)
        # Apply to A2:H{max_row}
        range_str = f"A2:H{max_row}"
        
        # Formula uses absolute column $G so entire row checks G's value
        ws.conditional_formatting.add(range_str, FormulaRule(formula=['$G2="Pass"'], stopIfTrue=True, fill=pass_fill))
        ws.conditional_formatting.add(range_str, FormulaRule(formula=['$G2="Fail"'], stopIfTrue=True, fill=fail_fill))
        ws.conditional_formatting.add(range_str, FormulaRule(formula=['$G2="Untested"'], stopIfTrue=True, fill=untested_fill))
        ws.conditional_formatting.add(range_str, FormulaRule(formula=['$G2="Skip (N/A)"'], stopIfTrue=True, fill=skip_fill))

    wb.save(output_path)
    print(f"Excel file created successfully at: {output_path}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_testcase.py <input_csv> <output_xlsx>")
        sys.exit(1)
    generate_excel(sys.argv[1], sys.argv[2])
